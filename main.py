import requests
import pandas as pd
import openpyxl
import time
from openpyxl.utils.dataframe import dataframe_to_rows


def fetch_live_data():
    url = "https://api.coingecko.com/api/v3/coins/markets"
    params = {
        'vs_currency': 'usd',
        'order': 'market_cap_desc',
        'per_page': 50,
        'page': 1,
        'sparkline': 'false'
    }
    response = requests.get(url, params=params)
    data = response.json()

    
    cleaned_data = []
    for item in data:
        cleaned_data.append({
            'name': item['name'],
            'symbol': item['symbol'],
            'current_price': item['current_price'],
            'market_cap': item['market_cap'],
            'total_volume': item['total_volume'],
            'price_change_percentage_24h': item['price_change_percentage_24h']
        })
    
    return pd.DataFrame(cleaned_data)

def analyze_data(df):
  
    top_5 = df.nlargest(5, 'market_cap')
  
    avg_price = df['current_price'].mean()
    
   
    highest_24h_change = df['price_change_percentage_24h'].max()
    lowest_24h_change = df['price_change_percentage_24h'].min()
    
    analysis = {
        'top_5': top_5,
        'avg_price': avg_price,
        'highest_24h_change': highest_24h_change,
        'lowest_24h_change': lowest_24h_change
    }
    
    return analysis


def update_excel(df, analysis, filename='crypto_data.xlsx'):
 
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Name', 'Symbol', 'Current Price (USD)', 'Market Cap', '24h Volume', '24h Price Change (%)'])
    
  
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None
    
   
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
   
    wb.save(filename)

def main():
    while True:
     
        df = fetch_live_data()
        
      
        print(df.head())
 
        analysis = analyze_data(df)
 
        print(analysis)
       
        update_excel(df, analysis)
      
        time.sleep(300)

if __name__ == "__main__":
    main()
